"""Local question runner for Tutor self-evaluation simulation."""

from __future__ import annotations

import json
import re
import shutil
from functools import lru_cache
from pathlib import Path
from typing import Any

from tools.constants import (
    CLOUD_SERVICES_ACTIVE,
    EXAMINER_SCORING_ALLOWED,
    KNOWLEDGE_DIR,
    PROJECT_ROOT,
    SAFE_FOR_EXAMINER,
    SELF_EVAL_DIR,
    USES_API,
    USES_EMBEDDINGS,
    USES_LLM,
    USES_VECTOR_DB,
)
from tools.orchestrator.orchestrator import run_orchestrator
from tools.self_eval.answer_comparator import compare_answer
from tools.self_eval.evaluation_reporter import DEFAULT_SELF_EVAL_DIR, write_evaluation_reports
from tools.tutor.answer_builder import build_tutor_answer


QUESTION_BANK_ROOT = KNOWLEDGE_DIR / "question-bank"
DEFAULT_QUESTION_LIMIT = 10
SAMPLE_QUESTIONS_PATH = SELF_EVAL_DIR / "sample_questions.json"
QUESTION_EXPECTATIONS_PATH = SELF_EVAL_DIR / "question_expectations.json"


def run_self_eval(
    question_type: str = "all",
    limit: int = DEFAULT_QUESTION_LIMIT,
    output_dir: Path = DEFAULT_SELF_EVAL_DIR,
    strictness: str = "hard",
) -> dict[str, Any]:
    """Run internal questions through the local Tutor pipeline."""
    questions = load_questions(question_type=question_type, limit=limit)
    output_dir.mkdir(parents=True, exist_ok=True)
    attempts_dir = output_dir / "attempts"
    results = []
    for question in questions:
        result = run_question_attempt(question, attempts_dir, strictness=strictness)
        results.append(result)
    report_paths = write_evaluation_reports(
        results,
        output_dir=output_dir,
        strictness=strictness,
        memory_path=output_dir / "pedagogical_memory.json",
    )
    return {
        "questions_attempted": len(results),
        "results": results,
        "report_paths": report_paths,
        "governance": {
            "safe_for_examiner": SAFE_FOR_EXAMINER,
            "examiner_scoring_allowed": EXAMINER_SCORING_ALLOWED,
            "uses_llm": USES_LLM,
            "uses_api": USES_API,
            "uses_embeddings": USES_EMBEDDINGS,
            "uses_vector_db": USES_VECTOR_DB,
            "cloud_services_active": CLOUD_SERVICES_ACTIVE,
            "strictness": strictness,
        },
    }


def run_question_attempt(question: dict[str, Any], attempts_dir: Path, strictness: str = "hard") -> dict[str, Any]:
    """Run one question through orchestrator, retrieval, Tutor synthesis, and comparator."""
    _enforce_question_governance(question)
    question_id = _safe_id(str(question.get("question_id") or "question"))
    question_dir = attempts_dir / question_id
    question_dir.mkdir(parents=True, exist_ok=True)
    orchestrator_result = run_orchestrator(
        query=str(question.get("question_text", "")),
        top_k=5,
        language="es",
        context_package_dir=question_dir,
        root=PROJECT_ROOT,
    )
    context_path = Path(orchestrator_result["context_package_paths"]["latest"])
    saved_context_path = question_dir / "tutor_context_package.json"
    shutil.copyfile(context_path, saved_context_path)
    tutor_result = build_tutor_answer(
        context_package_path=saved_context_path,
        output_path=question_dir / "tutor_attempt.md",
        style="standard",
    )
    answer_text = Path(tutor_result["output_paths"]["latest"]).read_text(encoding="utf-8")
    context_package = json.loads(saved_context_path.read_text(encoding="utf-8"))
    comparison = compare_answer(question, answer_text, context_package, strictness=strictness)
    result = {
        "question_id": question.get("question_id", ""),
        "question_type": question.get("question_type", "theory"),
        "question_text": question.get("question_text", ""),
        "expected_topics": question.get("expected_topics", []),
        "expected_causal_links": question.get("expected_causal_links", []),
        "expected_keywords": question.get("expected_keywords", []),
        "expected_reasoning_type": question.get("expected_reasoning_type", ""),
        "difficulty": question.get("difficulty", "intermediate"),
        "source_type": question.get("source_type", "internal_adapter"),
        "safe_for_examiner": SAFE_FOR_EXAMINER,
        "tutor_attempt_path": tutor_result["output_paths"]["latest"],
        "tutor_context_package_path": saved_context_path.as_posix(),
        "comparison": comparison,
        "governance": {
            "safe_for_examiner": SAFE_FOR_EXAMINER,
            "examiner_scoring_allowed": EXAMINER_SCORING_ALLOWED,
            "official_marks_assigned": False,
            "strictness": strictness,
        },
    }
    (question_dir / "self_eval_result.json").write_text(json.dumps(result, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    return result


def load_questions(
    question_type: str = "all",
    limit: int = DEFAULT_QUESTION_LIMIT,
    question_bank_root: Path = QUESTION_BANK_ROOT,
) -> list[dict[str, Any]]:
    """Load internal question-bank questions through a non-destructive adapter."""
    questions = _load_structured_questions(question_bank_root / "structured")
    if not questions:
        questions = _load_xlsx_questions(question_bank_root / "raw")
    if not questions:
        questions = _sample_questions()
    normalized = [_normalize_question(question, index) for index, question in enumerate(questions, start=1)]
    if question_type != "all":
        normalized = [question for question in normalized if question.get("question_type") == question_type]
    return normalized[: max(0, limit)]


def _load_structured_questions(path: Path) -> list[dict[str, Any]]:
    questions = []
    if not path.exists():
        return questions
    for file_path in sorted(path.glob("*.json")):
        data = json.loads(file_path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            questions.extend(data)
        elif isinstance(data, dict):
            questions.append(data)
    for file_path in sorted(path.glob("*.jsonl")):
        for line in file_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                questions.append(json.loads(line))
    return questions


def _load_xlsx_questions(path: Path) -> list[dict[str, Any]]:
    workbook_paths = sorted(path.glob("*.xlsx"))
    if not workbook_paths:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    questions = []
    workbook = load_workbook(workbook_paths[0], read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "").strip().lower() for value in rows[0]]
        for row_index, row in enumerate(rows[1:], start=2):
            record = {headers[index]: row[index] for index in range(min(len(headers), len(row))) if headers[index]}
            text = record.get("question") or record.get("question_text") or record.get("pregunta") or record.get("stem")
            if not text:
                continue
            questions.append(
                {
                    "question_id": record.get("question_id") or record.get("id") or f"{sheet.title}_{row_index}",
                    "question_text": str(text),
                    "question_type": _infer_question_type(str(text)),
                    "expected_topics": _split_field(record.get("topic") or record.get("topics") or record.get("expected_topics")),
                    "expected_causal_links": _split_field(record.get("expected_causal_links")),
                    "expected_keywords": _split_field(record.get("keywords") or record.get("expected_keywords")),
                    "expected_reasoning_type": str(record.get("reasoning_type") or record.get("expected_reasoning_type") or ""),
            "source_type": "raw_question_bank_adapter",
            "safe_for_examiner": False,
                }
            )
            if len(questions) >= 100:
                return questions
    return questions


def _normalize_question(question: dict[str, Any], index: int) -> dict[str, Any]:
    text = str(question.get("question_text") or question.get("question") or "")
    inferred = _infer_expectations(text)
    question_type = str(question.get("question_type") or _infer_question_type(text)).lower()
    if question_type not in {"theory", "sat", "short_answer"}:
        question_type = "sat" if question_type == "tasting" else "theory"
    return {
        "question_id": str(question.get("question_id") or f"SELF_EVAL_{index:03d}"),
        "question_text": text,
        "question_type": question_type,
        "expected_topics": _as_list(question.get("expected_topics")) or inferred["expected_topics"],
        "expected_causal_links": _as_list(question.get("expected_causal_links")) or inferred["expected_causal_links"],
        "expected_keywords": _as_list(question.get("expected_keywords")) or inferred["expected_keywords"],
        "expected_reasoning_type": str(question.get("expected_reasoning_type") or _infer_reasoning_type(text)),
        "difficulty": str(question.get("difficulty") or inferred["difficulty"]),
        "source_type": str(question.get("source_type") or "internal_adapter"),
        "safe_for_examiner": False,
    }


def _sample_questions() -> list[dict[str, Any]]:
    return list(_load_sample_questions())


@lru_cache(maxsize=1)
def _load_sample_questions() -> tuple[dict[str, Any], ...]:
    return tuple(json.loads(SAMPLE_QUESTIONS_PATH.read_text(encoding="utf-8")))


@lru_cache(maxsize=1)
def _load_question_expectations() -> tuple[dict[str, Any], ...]:
    return tuple(json.loads(QUESTION_EXPECTATIONS_PATH.read_text(encoding="utf-8")))


def _infer_question_type(text: str) -> str:
    lowered = text.lower()
    if "sat" in lowered or "tasting" in lowered or "quality assessment" in lowered:
        return "sat"
    if len(text.split()) < 10:
        return "short_answer"
    return "theory"


def _infer_reasoning_type(text: str) -> str:
    lowered = text.lower()
    if "how" in lowered or "affect" in lowered or "explain" in lowered:
        return "cause_effect"
    if "sat" in lowered or "quality assessment" in lowered:
        return "sat_logic"
    return "theory_foundation"


def _infer_expectations(text: str) -> dict[str, Any]:
    lowered = text.lower()
    for template in _load_question_expectations():
        if re.search(str(template.get("topic_pattern", "")), lowered):
            return {
                "expected_topics": _as_list(template.get("expected_topics")),
                "expected_causal_links": _as_list(template.get("expected_causal_links")),
                "expected_keywords": _as_list(template.get("expected_keywords")),
                "difficulty": str(template.get("difficulty") or "intermediate"),
            }
    return {
        "expected_topics": [],
        "expected_causal_links": [],
        "expected_keywords": [],
        "difficulty": "intermediate",
    }


def _split_field(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).replace("|", ";").split(";") if item.strip()]


def _as_list(value: Any) -> list[str]:
    return _split_field(value)


def _safe_id(value: str) -> str:
    return "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in value)[:80]


def _enforce_question_governance(question: dict[str, Any]) -> None:
    if question.get("safe_for_examiner"):
        raise ValueError("Self-eval questions must keep safe_for_examiner false.")

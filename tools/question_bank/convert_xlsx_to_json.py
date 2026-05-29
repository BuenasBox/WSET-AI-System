"""Convert WSET3 question bank Excel -> structured JSON for load_questions()."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

try:
    from tools.constants import KNOWLEDGE_DIR
except ImportError:
    KNOWLEDGE_DIR = Path(__file__).parent.parent.parent / "knowledge"

XLSX_PATH = KNOWLEDGE_DIR / "question-bank" / "raw" / "WSET3_Banco_Maestro_V9.xlsx"
STRUCTURED_DIR = KNOWLEDGE_DIR / "question-bank" / "structured"
SHEET_NAME = "Banco_Completo"
OUTPUT_FILENAME = "wset3_questions.json"
SAFE_OUTPUT_FILENAME = "wset3_questions.generated.json"

# IDs present in the manually curated original wset3_questions.json (1–25)
_KNOWN_ORIGINAL_IDS: frozenset[str] = frozenset(str(i) for i in range(1, 26))

# Excel Tipo -> schema question_type; unrecognized values produce None (skipped).
# "Abierta" (open-ended) maps to short_answer. These questions have no Respuesta_Texto
# or MCQ options by design — they are evaluator-agent training prompts, not MCQ items.
# Empty expected_keywords for Abierta is valid and must never be treated as an error.
TIPO_MAP: dict[str, str] = {
    "opcion_multiple": "theory",
    "verdadero_falso": "theory",
    "emparejamiento": "theory",
    "abierta": "short_answer",
}

# Tipo values that legitimately carry no Respuesta_Texto/opciones.
# Keyword-warning suppression applies to these.
_TIPOS_WITHOUT_ANSWER_TEXT: frozenset[str] = frozenset({"abierta"})

_LETTER_TO_OPTION: dict[str, str] = {
    "A": "Opcion_A",
    "B": "Opcion_B",
    "C": "Opcion_C",
    "D": "Opcion_D",
}


# ---------------------------------------------------------------------------
# Pure helper functions
# ---------------------------------------------------------------------------

def _is_safe_to_overwrite(existing: Any) -> bool:
    """True only when existing JSON is exactly the 25 known original entries."""
    if not isinstance(existing, list) or len(existing) != 25:
        return False
    ids = {str(entry.get("question_id", "")) for entry in existing}
    return ids == _KNOWN_ORIGINAL_IDS


def _has_manual_metadata(entry: dict[str, Any]) -> bool:
    """True if entry has at least one non-empty curated array."""
    return bool(
        entry.get("expected_topics")
        or entry.get("expected_causal_links")
        or entry.get("expected_keywords")
    )


def _map_tipo(tipo: str) -> str | None:
    """Map Excel Tipo to schema question_type; None if unrecognized."""
    return TIPO_MAP.get(tipo.strip().lower())


def _derive_keywords(row: dict[str, Any]) -> tuple[list[str], bool]:
    """
    Return (keywords, has_warning).

    Priority: Respuesta_Texto > option text for Respuesta_Letra > empty.
    has_warning=True only when neither source produced content.
    """
    texto = str(row.get("Respuesta_Texto") or "").strip()
    if texto:
        return [texto], False

    letra = str(row.get("Respuesta_Letra") or "").strip().upper()
    if letra in _LETTER_TO_OPTION:
        option_text = str(row.get(_LETTER_TO_OPTION[letra]) or "").strip()
        if option_text:
            return [option_text], False

    return [], True


def _build_expected_topics(row: dict[str, Any]) -> list[str]:
    ra = str(row.get("RA") or "").strip()
    bloque = str(row.get("Bloque") or "").strip()
    if ra and bloque:
        return [f"{ra} / {bloque}"]
    if ra:
        return [ra]
    if bloque:
        return [bloque]
    return []


def _build_question_from_row(row: dict[str, Any]) -> dict[str, Any]:
    """Generate a full structured question entry from an Excel row dict."""
    keywords, _ = _derive_keywords(row)
    return {
        "question_id": str(row.get("ID") or ""),
        "question_text": str(row.get("Pregunta") or ""),
        "question_type": _map_tipo(str(row.get("Tipo") or "")),
        "expected_topics": _build_expected_topics(row),
        "expected_causal_links": [],
        "expected_keywords": keywords,
        "expected_reasoning_type": "theory_foundation",
        "difficulty": "intermediate",
        "source_type": str(row.get("Fuente") or "excel_question_bank"),
        "safe_for_examiner": False,
        # Extra fields — not consumed by _normalize_question; preserved for future use
        "options": {
            "A": str(row.get("Opcion_A") or ""),
            "B": str(row.get("Opcion_B") or ""),
            "C": str(row.get("Opcion_C") or ""),
            "D": str(row.get("Opcion_D") or ""),
        },
        "correct_answer_letter": str(row.get("Respuesta_Letra") or ""),
        "correct_answer_text": str(row.get("Respuesta_Texto") or ""),
    }


def _merge_preserved_with_extras(
    preserved: dict[str, Any], row: dict[str, Any]
) -> dict[str, Any]:
    """Keep all curated schema fields from preserved; add MCQ extras from Excel."""
    entry = dict(preserved)
    entry["options"] = {
        "A": str(row.get("Opcion_A") or ""),
        "B": str(row.get("Opcion_B") or ""),
        "C": str(row.get("Opcion_C") or ""),
        "D": str(row.get("Opcion_D") or ""),
    }
    entry["correct_answer_letter"] = str(row.get("Respuesta_Letra") or "")
    entry["correct_answer_text"] = str(row.get("Respuesta_Texto") or "")
    return entry


# ---------------------------------------------------------------------------
# Main converter
# ---------------------------------------------------------------------------

def convert(
    xlsx_path: Path = XLSX_PATH,
    structured_dir: Path = STRUCTURED_DIR,
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """
    Convert WSET3 question bank Excel to structured JSON.

    Returns a stats dict:
        excel_rows      — total data rows read from Banco_Completo
        converted       — rows generated fresh from Excel
        preserved       — rows whose manual metadata was kept from existing JSON
        skipped         — invalid rows not written (see skipped_details)
        warnings        — rows written but with no derivable keyword
        skipped_details — list of {row, id, reason} for each skipped row
        warning_details — list of {row, id, reason} for each warning
        output_path     — absolute path of the JSON file written
        safe_overwrite  — True if primary path was used, False if safe path used
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    output_path = structured_dir / OUTPUT_FILENAME
    safe_output_path = structured_dir / SAFE_OUTPUT_FILENAME

    # Determine whether we can safely overwrite the primary output
    existing_entries: list[dict[str, Any]] = []
    if not output_path.exists():
        safe_to_overwrite = True
    else:
        try:
            existing_entries = json.loads(output_path.read_text(encoding="utf-8"))
            safe_to_overwrite = _is_safe_to_overwrite(existing_entries)
        except (json.JSONDecodeError, UnicodeDecodeError):
            existing_entries = []
            safe_to_overwrite = False

    # Build the set of manually curated entries to preserve
    preserved_manual: dict[str, dict[str, Any]] = {
        str(entry["question_id"]): entry
        for entry in existing_entries
        if _has_manual_metadata(entry)
    }

    # Read the Excel (read-only, data-only — never modifies the source)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {xlsx_path.name}")
    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError(f"Sheet '{SHEET_NAME}' is empty.")

    headers = [str(v or "").strip() for v in all_rows[0]]
    data_rows = all_rows[1:]

    stats: dict[str, Any] = {
        "excel_rows": len(data_rows),
        "converted": 0,
        "preserved": 0,
        "skipped": 0,
        "warnings": 0,
        "skipped_details": [],
        "warning_details": [],
        "output_path": None,
        "safe_overwrite": safe_to_overwrite,
    }

    result: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for row_idx, raw_row in enumerate(data_rows, start=2):
        row: dict[str, Any] = {
            headers[i]: raw_row[i]
            for i in range(min(len(headers), len(raw_row)))
            if headers[i]
        }

        qid = str(row.get("ID") or "").strip()
        pregunta = str(row.get("Pregunta") or "").strip()
        tipo_raw = str(row.get("Tipo") or "").strip()

        if not pregunta:
            stats["skipped"] += 1
            stats["skipped_details"].append(
                {"row": row_idx, "id": qid, "reason": "empty Pregunta"}
            )
            continue

        if qid in seen_ids:
            stats["skipped"] += 1
            stats["skipped_details"].append(
                {"row": row_idx, "id": qid, "reason": "duplicate ID"}
            )
            continue
        seen_ids.add(qid)

        question_type = _map_tipo(tipo_raw)
        if question_type is None:
            stats["skipped"] += 1
            stats["skipped_details"].append(
                {
                    "row": row_idx,
                    "id": qid,
                    "reason": f"unrecognized Tipo: '{tipo_raw}'",
                }
            )
            continue

        _, has_keyword_warning = _derive_keywords(row)
        # Suppress the keyword warning for Abierta (open-ended) questions: they
        # intentionally have no Respuesta_Texto or MCQ options. An empty keyword
        # list is valid for this type and must not invite filtering or cleanup.
        if has_keyword_warning and tipo_raw.lower() not in _TIPOS_WITHOUT_ANSWER_TEXT:
            stats["warnings"] += 1
            stats["warning_details"].append(
                {
                    "row": row_idx,
                    "id": qid,
                    "reason": "no Respuesta_Texto or derivable keyword from correct option",
                }
            )

        if qid in preserved_manual:
            entry = _merge_preserved_with_extras(preserved_manual[qid], row)
            stats["preserved"] += 1
        else:
            entry = _build_question_from_row(row)
            stats["converted"] += 1

        result.append(entry)

    def _sort_key(q: dict[str, Any]) -> tuple[int, str]:
        try:
            return (int(str(q["question_id"])), "")
        except (ValueError, KeyError):
            return (999_999, str(q.get("question_id", "")))

    result.sort(key=_sort_key)

    final_output_path = output_path if safe_to_overwrite else safe_output_path
    stats["output_path"] = str(final_output_path)

    if not dry_run:
        structured_dir.mkdir(parents=True, exist_ok=True)
        final_output_path.write_text(
            json.dumps(result, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )

    return stats


# ---------------------------------------------------------------------------
# CLI report
# ---------------------------------------------------------------------------

def _print_report(stats: dict[str, Any]) -> None:
    sep = "=" * 62
    print(f"\n{sep}")
    print("WSET3 Question Bank Converter — Report")
    print(sep)
    print(f"Excel rows read         : {stats['excel_rows']}")
    print(f"  Preserved (manual)    : {stats['preserved']}")
    print(f"  Converted (generated) : {stats['converted']}")
    print(f"  Skipped (invalid)     : {stats['skipped']}")
    print(f"  Warnings (no keyword) : {stats['warnings']}")
    cuadre = stats["preserved"] + stats["converted"] + stats["skipped"]
    status = "OK" if cuadre == stats["excel_rows"] else "MISMATCH"
    print(f"Cuadre check            : {cuadre} / {stats['excel_rows']} ({status})")
    if stats["skipped_details"]:
        print("\nSkipped rows:")
        for d in stats["skipped_details"]:
            print(f"  Row {d['row']:4d} | ID={str(d['id']):>6s} | {d['reason']}")
    if stats["warning_details"]:
        print("\nWarnings (no keyword derived):")
        for d in stats["warning_details"]:
            print(f"  Row {d['row']:4d} | ID={str(d['id']):>6s} | {d['reason']}")
    print(f"\nJSON written to : {stats['output_path']}")
    if not stats["safe_overwrite"]:
        print(
            "\n*** WRITTEN TO SAFE PATH — existing JSON had unexpected content. ***\n"
            "*** Review wset3_questions.generated.json and merge manually.     ***"
        )
    print(f"{sep}\n")


if __name__ == "__main__":
    stats = convert()
    _print_report(stats)
    if not stats["safe_overwrite"]:
        sys.exit(1)
